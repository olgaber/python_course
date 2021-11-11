"""
Файл имеет вид таблицы: Фамилия Имя Отдел Зарплата (в первой строке заголовки
колонок)
 Посчитайте сколько отделов на фирме
 Определите максимальную зарплату
 Определите максимальную зарплату в каждом отделе
 Выведите «Отдел Макс_Зарплата Фамилия_человека_с_такой_зарплатой» в
новый файл
Подсказка: используйте словари!!!
"""

from pathlib import Path
from collections import Counter
import openpyxl

file = Path("table.xlsx")
wb_obj = openpyxl.load_workbook(file)
sheet = wb_obj.active

keys = []
for row in sheet.iter_rows(max_row=1):
    for cell in row:
        keys.append(cell.value)
# print(keys)

dict_list = []
for r in range(2, 10):
    row_dict = {}
    for c in range(1, 5):
        value = sheet.cell(r, c).value
        row_dict[keys[c - 1]] = value
    dict_list.append(row_dict)
    # print(row_dict)


def create_departments_list(d_ls):
    dep_list = []
    for d in d_ls:
        dep_list.append(d["Department"])
    dep_list = list(set(dep_list))
    print(dep_list)
    return dep_list


def calculate_departments(d_ls):
    dep_list = create_departments_list(d_ls)
    employees_by_dept = dict(Counter(dep_list))
    print(employees_by_dept)
    print("1. Departments quantity: " + str(len(employees_by_dept)))


def max_company_salary(d_ls):
    salary_list = []
    for s in d_ls:
        salary_list.append(s["Salary"])
    print(salary_list)
    print("2. Max company salary: " + str(max(salary_list)))


def max_department_salary(d_ls):
    dep_list = create_departments_list(d_ls)
    dep_salary_list = []
    i = 0
    for dep in dep_list:
        for dic in d_ls:
            if dep == dic.get("Department"):
                dep_salary_list.insert(i, dic.get("Salary"))
        print(dep_salary_list)
        maximum = max(dep_salary_list)
        print(f"3. Max salary for {dep} department = {maximum}")
        i += 0
        dep_salary_list = []
        continue


if __name__ == '__main__':
    calculate_departments(dict_list)
    max_company_salary(dict_list)
    max_department_salary(dict_list)

